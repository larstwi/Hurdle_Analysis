#!/usr/bin/env python3
"""Personalisiertes Auswertungsblatt fuer einen 400-m-Huerden-Athleten.

Aufbau pro Rennen: zwei Zeilen uebereinander.
  obere Zeile  - Abschnittszeiten in Sekunden
  untere Zeile - Schrittzahl im selben Abschnitt, exakt darunter

Farben werden direkt auf die Zellen geschrieben (nicht als bedingte
Formatierung), damit sie auch in Numbers und in Vorschauen erhalten bleiben.
Alle Zellwerte sind fertige, in Python berechnete Werte statt Formeln auf
das Blatt "Rohdaten" - weder Excel-Formeln noch injizierte Formel-Caches
werden von der iOS-"Vorschau" (Quick Look) zuverlaessig ausgewertet.
"""

import sys
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef

from master_io import load_master
from auswertung import select_season, kuerzel_runde
from xlsx_cache import setze_rahmen_und_fuellung, stelle_apply_flags_sicher

ARIAL = 'Arial'
TINTE, GRAU = '1F3348', '6B7A8A'
RAND = 'B8C4D0'

# Schrittzeile wird durchgehend leicht grau hinterlegt, damit sie sich
# klar von der weissen Zeitzeile darueber abhebt
SCHRITTZEILE = 'EDF1F5'
# Persoenliche Bestzeit bekommt einen goldenen Rahmen
GOLD = 'B8860B'
GOLD_HELL = 'FBF0CE'

# Ein Abschnitt, der deutlich schneller ist als der vorangehende, widerspricht
# dem Ermuedungsverlauf und deutet auf einen Tippfehler im Touchdown hin.
duenn = Side(style='thin', color=RAND)
gold = Side(style='medium', color=GOLD)
RAHMEN = Border(left=duenn, right=duenn, top=duenn, bottom=duenn)
OBEN = Border(left=duenn, right=duenn, top=duenn)
UNTEN = Border(left=duenn, right=duenn, bottom=duenn)

KOPF_RENNEN = ['Datum', 'Ort', 'Rd', 'Bahn', 'Rang']
KOPF_ERGEBNIS = ['Zeit', '0–200', '200–400', 'Diff']
KOPF_SEGMENT = ['Start–H1'] + [f'H{i}–H{i+1}' for i in range(1, 10)] + ['H10–Ziel']

C_ZEIT, C_H200, C_H400, C_DIFF = 6, 7, 8, 9
C_SEG0 = 10                      # Start–H1
C_SEG35_0, C_SEG35_1 = 11, 19    # die neun 35-m-Abschnitte
C_SEGZ = 20                      # H10–Ziel
BREIT = C_SEGZ

BREITEN = [11, 17, 5, 6, 6, 9, 8, 9, 8, 10] + [13] * 9 + [14]

# Spalten auf "Rohdaten"
R_DATUM, R_ORT, R_RUNDE, R_LAUF, R_BAHN, R_RANG, R_ZEIT, R_STATUS = 2, 3, 4, 5, 6, 7, 8, 9
R_H1, R_H5, R_H6, R_H10, R_S0 = 10, 14, 15, 19, 20
R_LABEL, R_KURZ = 30, 31


def rd(col, zeile):
    return f'Rohdaten!{L(col)}{zeile}'


def kopfzelle(ws, zelle, text, groesse=11, fett=True, vordergrund='FFFFFF',
              fuellung=TINTE, ausrichtung='center'):
    c = ws[zelle]
    c.value = text
    c.font = Font(name=ARIAL, size=groesse, bold=fett, color=vordergrund)
    if fuellung:
        c.fill = PatternFill('solid', fgColor=fuellung)
    c.alignment = Alignment(horizontal=ausrichtung, vertical='center', wrap_text=True)
    return c


def schreibe_rohdaten(ws, df):
    import datetime as _dt
    kopf = (['race_id', 'datum', 'ort', 'runde', 'lauf', 'bahn', 'rang', 'zeit', 'status']
            + [f'h{i}' for i in range(1, 11)]
            + ['s_start'] + [f's{i}_{i+1}' for i in range(1, 10)])
    for j, h in enumerate(kopf, 1):
        kopfzelle(ws, f'{L(j)}1', h, groesse=9)
    for i, (_, r) in enumerate(df.iterrows(), start=2):
        for j, h in enumerate(kopf, 1):
            v = r[h]
            if v is None or (isinstance(v, float) and pd.isna(v)):
                v = ''
            if h == 'datum' and v:
                v = _dt.date.fromisoformat(str(v))
            c = ws.cell(i, j, v)
            c.font = Font(name=ARIAL, size=9)
            if h == 'datum':
                c.number_format = 'DD.MM.YYYY'
    kopfzelle(ws, f'{L(R_KURZ)}1', 'kurz', groesse=9)
    for i, (_, r) in enumerate(df.iterrows(), start=2):
        c = ws.cell(i, R_KURZ, kuerzel_runde(r['runde'], r['lauf']))
        c.font = Font(name=ARIAL, size=9)
    ws.column_dimensions[L(R_KURZ)].width = 7
    ws.freeze_panes = 'B2'
    for j, b in enumerate([22, 11, 13, 17, 5, 5, 5, 7, 7] + [7] * 19, 1):
        ws.column_dimensions[L(j)].width = b
    ws.sheet_view.showGridLines = False


def werte(r):
    """Abschnittszeiten und Schrittzahlen als Zahlen, fuer die Einfaerbung."""
    h = [r.get(f'h{i}') for i in range(1, 11)]
    h = [None if (v is None or v == '' or pd.isna(v)) else float(v) for v in h]
    z = r.get('zeit')
    z = None if (z is None or z == '' or pd.isna(z)) else float(z)

    seg = [h[0]]
    for i in range(9):
        seg.append(None if h[i] is None or h[i + 1] is None else round(h[i + 1] - h[i], 3))
    seg.append(None if h[9] is None or z is None else round(z - h[9], 3))

    s = [r.get('s_start')] + [r.get(f's{i}_{i+1}') for i in range(1, 10)]
    s = [None if (v is None or v == '' or pd.isna(v)) else int(v) for v in s]
    return seg, s


def rennblock(ws, blattname, oben, daten, blass=False):
    """Schreibt ein Rennen als zwei Zeilen. Gibt die naechste freie Zeile zurueck.

    Alle Zellwerte werden direkt aus den Python-Daten geschrieben (keine
    Formeln auf "Rohdaten") - siehe Kommentar weiter unten.
    """
    unten = oben + 1
    seg, schritte = werte(daten)
    ton = '5A6B7C' if blass else '1A2430'

    def num(v):
        return None if (v is None or v == '' or (isinstance(v, float) and pd.isna(v))) else float(v)

    h = [num(daten.get(f'h{i}')) for i in range(1, 11)]
    zeit_num = num(daten.get('zeit'))
    status = daten.get('status') or 'OK'
    kurz_txt = kuerzel_runde(daten.get('runde'), daten.get('lauf'))
    datum_txt = daten.get('datum')

    # Alle sichtbaren Zellen dieser Zeile sind fertige Werte, keine Formeln
    # auf "Rohdaten". Grund: weder ein injizierter Formel-Cache noch eine
    # echte Formel wird von der iOS-"Vorschau" (Quick Look) zuverlaessig
    # ausgewertet oder angezeigt - dort erscheint praktisch jede Formelzelle
    # als 0, unabhaengig vom Cache. Excel, LibreOffice und die vollwertige
    # Numbers-App rechnen zwar korrekt neu, aber Quick Look tut das nicht.
    # Ein fertig berechneter Wert ist unabhaengig von jeder Formel- oder
    # Neuberechnungs-Logik und damit ueberall garantiert korrekt sichtbar.
    # Datum zusaetzlich als fertig formatierter Text (siehe fruehere Notiz zu
    # Grossbuchstaben-Datumsformaten wie "DD.MM.YYYY").
    import datetime as _dt
    datum_str = ''
    if datum_txt:
        d = _dt.date.fromisoformat(str(datum_txt)) if isinstance(datum_txt, str) else datum_txt
        datum_str = d.strftime('%d.%m.%Y')
    ws.cell(oben, 1, datum_str)
    ws.cell(oben, 2, daten.get('ort') or '')
    ws.cell(oben, 3, kurz_txt or '')
    bahn = num(daten.get('bahn'))
    ws.cell(oben, 4, bahn if bahn is not None else '')
    rang = num(daten.get('rang'))
    ws.cell(oben, 5, rang if rang is not None else '')

    if status != 'OK':
        ws.cell(oben, C_ZEIT, status)
    else:
        ws.cell(oben, C_ZEIT, zeit_num if zeit_num is not None else '')

    m200 = None if h[4] is None or h[5] is None else h[4] + (h[5] - h[4]) * 14 / 35
    ws.cell(oben, C_H200, round(m200, 4) if m200 is not None else '')

    m400 = None if m200 is None or zeit_num is None else zeit_num - m200
    ws.cell(oben, C_H400, round(m400, 4) if m400 is not None else '')

    diff = None if m200 is None or m400 is None else m400 - m200
    ws.cell(oben, C_DIFF, round(diff, 4) if diff is not None else '')

    for j in range(1, C_DIFF + 1):
        ws.merge_cells(start_row=oben, start_column=j, end_row=unten, end_column=j)

    ws.cell(oben, C_SEG0, h[0] if h[0] is not None else '')

    for i in range(9):
        if h[i] is not None and h[i + 1] is not None:
            ws.cell(oben, C_SEG35_0 + i, f'{h[i+1]-h[i]:.2f} ({h[i+1]:.2f})')
        else:
            ws.cell(oben, C_SEG35_0 + i, '')

    if h[9] is not None and zeit_num is not None:
        ws.cell(oben, C_SEGZ, f'{zeit_num-h[9]:.2f} ({zeit_num:.2f})')
    else:
        ws.cell(oben, C_SEGZ, '')

    s_namen = ['s_start'] + [f's{k}_{k+1}' for k in range(1, 10)]
    for i in range(10):
        sv = daten.get(s_namen[i])
        sv = None if (sv in (None, '') or (isinstance(sv, float) and pd.isna(sv))) else int(float(sv))
        ws.cell(unten, C_SEG0 + i, sv if sv is not None else '')

    for j in range(1, BREIT + 1):
        c = ws.cell(oben, j)
        c.font = Font(name=ARIAL, size=10, color=ton, bold=(j == C_ZEIT), italic=blass)
        c.border = OBEN if j >= C_SEG0 else RAHMEN
        c.alignment = Alignment(horizontal='left' if j in (2, 3) else 'center',
                                vertical='center')
        if j == 1:
            c.number_format = '@'
        elif j == C_DIFF:
            c.number_format = '+0.00;−0.00;0.00'
        elif j == C_ZEIT or j in (C_H200, C_H400) or j == C_SEG0:
            c.number_format = '0.00'
        elif j > C_SEG0:
            c.number_format = '@'   # Text: "Segment (Zwischenzeit)"

    for j in range(1, C_DIFF + 1):
        ws.cell(unten, j).border = UNTEN

    for j in range(C_SEG0, BREIT + 1):
        c = ws.cell(unten, j)
        c.font = Font(name=ARIAL, size=9, color='3D4B59', italic=blass)
        c.border = UNTEN
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.number_format = '0'
        c.fill = PatternFill('solid', fgColor=SCHRITTZEILE)

    ws.row_dimensions[oben].height = 17
    ws.row_dimensions[unten].height = 13
    return unten + 1


def markiere_bestzeit(ws, oben, unten, breit):
    """Setzt die Kommentarmarkierung der PB und liefert die gewuenschten
    Rahmen/Fuellungen der PB-Zeile als Zellspezifikation zurueck (fuer
    xlsx_cache.setze_rahmen_und_fuellung), statt sie direkt ueber openpyxl
    zu setzen.

    Grund: bei so vielen ueber zwei Zeilen verschmolzenen Zellen im Blatt
    verwirft bzw. vertauscht openpyxl beim Speichern reproduzierbar den
    Rahmen einzelner verschmolzener Zellen - im Python-Objekt korrekt, in
    der gespeicherten Datei nicht mehr. baue() patcht die Werte darum nach
    dem Speichern direkt im XML, das ist zuverlaessig.

    Die Spalten 1-9 sind je Rennen ueber oben+unten zu einer sichtbaren Zelle
    verschmolzen (siehe rennblock). Excel bildet den sichtbaren Aussenrand
    einer verschmolzenen Zelle aus den Randzellen des Bereichs (oben liefert
    die obere Kante, unten die untere usw.), darum bekommen bei diesen
    Spalten beide Zellen (oben und unten) denselben vollstaendigen Rahmen.
    Bei den echten zweizeiligen Abschnittsspalten (10-BREIT) bleibt es beim
    einfachen Aufteilen (oben nur top, unten nur bottom) - Excel und die
    iOS-Vorschau zeigen das korrekt; nur die volle Numbers-App auf dem Mac
    zeigt dort noch keine untere Aussenkante. Eine "beidseitig volle"
    Variante wuerde das zwar kaschieren, erzeugt aber sichtbar eine
    zusaetzliche goldene Linie mitten durch die Abschnittsspalten - das ist
    schlechter als die verbleibende Numbers-Unschoenheit.
    """
    GOLD_SEITE = ('medium', '00' + GOLD)
    GRAU_SEITE = ('thin', '00' + RAND)
    zellen = {}
    for j in range(1, breit + 1):
        links = GOLD_SEITE if j == 1 else GRAU_SEITE
        rechts = GOLD_SEITE if j == breit else GRAU_SEITE
        fuellung = GOLD_HELL if j <= C_DIFF else None
        if j <= C_DIFF:
            for r in (oben, unten):
                zellen[f'{L(j)}{r}'] = dict(left=links, right=rechts,
                                             top=GOLD_SEITE, bottom=GOLD_SEITE, fill=fuellung)
        else:
            zellen[f'{L(j)}{oben}'] = dict(left=links, right=rechts, top=GOLD_SEITE, bottom=None)
            zellen[f'{L(j)}{unten}'] = dict(left=links, right=rechts, top=None, bottom=GOLD_SEITE)
    kopf = ws.cell(oben, 1)
    kopf.comment = Comment('Persönliche Bestzeit', 'Auswertung')
    return zellen


def baue(master, athlet, saison, ziel, vergleiche=4):
    auswahl = select_season(master, athlet, saison, vergleiche)
    lauf, vgl = auswahl['lauf'], auswahl['vgl']
    sb, pb, pb_jahr, pb_id = auswahl['sb'], auswahl['pb'], auswahl['pb_jahr'], auswahl['pb_id']

    reihenfolge = pd.concat([lauf, vgl])
    idx = {rid: i + 2 for i, rid in enumerate(reihenfolge['race_id'])}

    wb = Workbook()
    ws = wb.active
    ws.title = 'Saison'
    roh = wb.create_sheet('Rohdaten')
    schreibe_rohdaten(roh, reihenfolge)

    for j, b in enumerate(BREITEN, 1):
        ws.column_dimensions[L(j)].width = b

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=BREIT)
    kopfzelle(ws, 'A1', f'{str(athlet).upper()}   ·   400 M HÜRDEN   ·   SAISON {saison}',
              groesse=17, ausrichtung='left')
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18

    kpi = [('Saisonbestzeit', f'{sb:.2f} s' if pd.notna(sb) else '—'),
           ('Persönliche Bestzeit', f'{pb:.2f} s' if pd.notna(pb) else '—'),
           ('Jahr', str(pb_jahr) if pb_jahr else '—'),
           ('Rennen in der Saison', str(len(lauf))),
           ('Beendet', str(int((lauf['status'] == 'OK').sum())))]
    sp = 1
    for titel, wert in kpi:
        ws.merge_cells(start_row=3, start_column=sp, end_row=3, end_column=sp + 2)
        ws.merge_cells(start_row=4, start_column=sp, end_row=4, end_column=sp + 2)
        kopfzelle(ws, f'{L(sp)}3', titel, groesse=8, fett=False,
                  vordergrund=GRAU, fuellung=None, ausrichtung='left')
        kopfzelle(ws, f'{L(sp)}4', wert, groesse=14, vordergrund=TINTE,
                  fuellung=None, ausrichtung='left')
        sp += 3
    ws.row_dimensions[4].height = 20

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=BREIT)
    kopfzelle(ws, 'A5',
              'Je Rennen zwei Zeilen: oben die Abschnittszeit mit Zwischenzeit seit Start in '
              'Klammern, darunter die Schrittzahl im gleichen Abschnitt.   '
              'Goldener Rahmen = persönliche Bestzeit.',
              groesse=8, fett=False, vordergrund=GRAU, fuellung=None, ausrichtung='left')
    ws.row_dimensions[5].height = 14

    for a, b, t in [(1, 5, 'RENNEN'), (C_ZEIT, C_DIFF, 'ERGEBNIS'),
                    (C_SEG0, C_SEGZ, 'ABSCHNITTE   ·   Sekunden (Zwischenzeit)   ·   unten Schritte')]:
        ws.merge_cells(start_row=7, start_column=a, end_row=7, end_column=b)
        kopfzelle(ws, f'{L(a)}7', t, groesse=9)
    for j, t in enumerate(KOPF_RENNEN + KOPF_ERGEBNIS + KOPF_SEGMENT, 1):
        c = kopfzelle(ws, f'{L(j)}8', t, groesse=9, fuellung='DCE3EA', vordergrund=TINTE)
        c.border = RAHMEN
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[8].height = 24

    z = 9
    saison_zeilen, block_von = [], {}
    for _, r in lauf.iterrows():
        saison_zeilen.append(z)
        block_von[r['race_id']] = z
        z = rennblock(ws, 'Saison', z, r)

    if not vgl.empty:
        z += 1
        ws.merge_cells(start_row=z, start_column=1, end_row=z, end_column=BREIT)
        kopfzelle(ws, f'A{z}', 'VERGLEICH FRÜHERE JAHRE   ·   bestes Rennen je Saison, neuestes zuerst',
                  groesse=9, fuellung=GRAU, ausrichtung='left')
        z += 1
        for _, r in vgl.iterrows():
            block_von[r['race_id']] = z
            z = rennblock(ws, 'Saison', z, r, blass=True)
    letzte = z - 1

    pb_rahmen = {}
    if pb_id and pb_id in block_von:
        o = block_von[pb_id]
        pb_rahmen = markiere_bestzeit(ws, o, o + 1, BREIT)

    # ---------- Bezeichner fuer die Diagrammlegenden ----------
    # Auch hier bewusst als fertiger Text statt TEXT()-Verkettungsformel
    # geschrieben (gleicher Grund wie bei den Abschnittszellen oben) -
    # sonst zeigen die Diagrammlegenden in Numbers/Vorschau nichts an.
    kopfzelle(roh, f'{L(R_LABEL)}1', 'legende', groesse=9)
    roh.column_dimensions[L(R_LABEL)].width = 24
    label_von = {}
    for _, r in reihenfolge.iterrows():
        q = idx[r['race_id']]
        kurz = kuerzel_runde(r['runde'], r['lauf'])
        datum_txt = pd.to_datetime(r['datum']).strftime('%d.%m.') if r.get('datum') else ''
        label = f"{datum_txt} {r['ort']}" + (f' {kurz}' if kurz else ' ')
        label_von[r['race_id']] = label
        c = roh.cell(q, R_LABEL, label)
        c.font = Font(name=ARIAL, size=9)

    # ---------- Rechenblatt fuer die Kurven ----------
    # Blatt fuer die Diagramme, Rechnung liegt auf einem eigenen Blatt dahinter
    gr = wb.create_sheet('Grafiken')
    gr.sheet_view.showGridLines = False
    kv = wb.create_sheet('Berechnung')
    kv.sheet_view.showGridLines = False
    kv.column_dimensions['A'].width = 26
    for j in range(2, 14):
        kv.column_dimensions[L(j)].width = 9
    KOPFRAUM = 1

    def beschriftung(zeile, spalten, texte):
        for j, t in enumerate(texte, spalten):
            kopfzelle(kv, f'{L(j)}{zeile}', t, groesse=9)

    # --- Block 1: kumulierter Rueckstand zum Referenzrennen ---
    ref = auswahl['ref']
    ref_name = ''
    if ref is not None:
        ref_name = (f"{pd.to_datetime(ref['datum']).strftime('%d.%m.%y')} {ref['ort']}"
                    f" — {float(ref['zeit']):.2f} s")
    kopfzelle(kv, f'A{KOPFRAUM}', f'Kumulierter Rückstand zur Referenz ({ref_name})',
              groesse=10, ausrichtung='left')
    kv.merge_cells(f'A{KOPFRAUM}:L{KOPFRAUM}')
    kat1 = KOPFRAUM + 1
    beschriftung(kat1, 2, [f'H{i}' for i in range(1, 11)] + ['Ziel'])
    kopfzelle(kv, f'A{kat1}', 'Rennen', groesse=9, ausrichtung='left')

    z1 = kat1 + 1
    gap_zeilen = []
    if ref is not None:
        rq = idx[ref['race_id']]
        for _, r in lauf.iterrows():
            q = idx[r['race_id']]
            kv.cell(z1, 1, f'={rd(R_LABEL, q)}').font = Font(name=ARIAL, size=9)
            for i in range(10):
                a, b = rd(R_H1 + i, q), rd(R_H1 + i, rq)
                kv.cell(z1, 2 + i, f'=IF(OR({a}="",{b}=""),NA(),{a}-{b})')
            za, zb = rd(R_ZEIT, q), rd(R_ZEIT, rq)
            kv.cell(z1, 12, f'=IF(OR({za}="",{zb}=""),NA(),{za}-{zb})')
            for j in range(2, 13):
                kv.cell(z1, j).number_format = '+0.00;−0.00;0.00'
                kv.cell(z1, j).font = Font(name=ARIAL, size=9)
            gap_zeilen.append(z1)
            z1 += 1

    # --- Block 2: Ermuedungsprofil ---
    z2 = z1 + 2
    kopfzelle(kv, f'A{z2}', 'Ermüdungsprofil — Abschnitt minus schnellster Abschnitt '
                            'desselben Rennens (Sekunden)', groesse=10, ausrichtung='left')
    kv.merge_cells(f'A{z2}:L{z2}')
    z2 += 1
    beschriftung(z2, 2, [f'H{i}–H{i+1}' for i in range(1, 10)])
    kopfzelle(kv, f'A{z2}', 'Rennen', groesse=9, ausrichtung='left')
    kopf2 = z2
    z2 += 1

    # Hilfsspalten N..V: die neun Abschnitte je Rennen
    HILF = 14
    kopfzelle(kv, f'{L(HILF)}{kopf2}', 'Abschnitte (Hilfsspalten)', groesse=8)
    kv.merge_cells(start_row=kopf2, start_column=HILF, end_row=kopf2, end_column=HILF + 8)

    prof_zeilen = []
    for _, r in reihenfolge.iterrows():
        q = idx[r['race_id']]
        kv.cell(z2, 1, f'={rd(R_LABEL, q)}').font = Font(name=ARIAL, size=9)
        for i in range(9):
            a, b = rd(R_H1 + i, q), rd(R_H1 + i + 1, q)
            kv.cell(z2, HILF + i, f'=IF(OR({a}="",{b}=""),"",{b}-{a})')
            kv.cell(z2, HILF + i).font = Font(name=ARIAL, size=8, color=GRAU)
        bereich = f'{L(HILF)}{z2}:{L(HILF + 8)}{z2}'
        for i in range(9):
            zelle = f'{L(HILF + i)}{z2}'
            kv.cell(z2, 2 + i,
                    f'=IF({zelle}="",NA(),{zelle}-MIN({bereich}))')
            kv.cell(z2, 2 + i).number_format = '0.00'
            kv.cell(z2, 2 + i).font = Font(name=ARIAL, size=9)
        prof_zeilen.append((z2, r['race_id']))
        z2 += 1
    for j in range(HILF, HILF + 9):
        kv.column_dimensions[L(j)].width = 8

    # ---------- Diagramme ----------
    # Serienbeschriftungen liegen auf dem sichtbaren Grafikblatt, sonst
    # zeigt die Legende beim Rendern ins Leere
    lbl_sp = 16          # Spalte P, ausserhalb des Druckbereichs
    lbl_lauf = [1]

    def beschrifte(race_ids):
        zeilen = []
        for rid in race_ids:
            r0 = lbl_lauf[0]
            gr.cell(r0, lbl_sp, label_von.get(rid, '')).font = Font(name=ARIAL, size=9)
            zeilen.append(r0)
            lbl_lauf[0] += 1
        gr.column_dimensions[L(lbl_sp)].width = 24
        return zeilen

    def linie(titel, y_titel, daten_zeilen, kat_zeile, min_sp, max_sp,
              lbl_zeilen, hoehe=8.2, breite=25):
        ch = LineChart()
        ch.title = titel
        ch.style = 2
        ch.y_axis.title = y_titel
        ch.height, ch.width = hoehe, breite
        ch.y_axis.majorGridlines = None
        ch.legend.position = 'b'
        for r0 in daten_zeilen:
            ch.add_data(Reference(kv, min_col=min_sp, max_col=max_sp, min_row=r0, max_row=r0),
                        from_rows=True, titles_from_data=False)
        ch.set_categories(Reference(kv, min_col=min_sp, max_col=max_sp,
                                    min_row=kat_zeile, max_row=kat_zeile))
        for k, se in enumerate(ch.series):
            se.tx = SeriesLabel(strRef=StrRef(f'Grafiken!${L(lbl_sp)}${lbl_zeilen[k]}'))
            se.smooth = False
        return ch

    if gap_zeilen:
        c1 = linie(f'Wo wird die Zeit gewonnen und verloren?   Referenz: {ref_name}'
                   '   ·   unter der Nulllinie = schneller',
                   'Sekunden', gap_zeilen, kat1, 2, 12,
                   beschrifte(list(lauf['race_id'])), hoehe=8.6, breite=26)
        c1.x_axis.title = 'Hürde'
        gr.add_chart(c1, 'A1')

    if prof_zeilen:
        c2 = linie('Ermüdungsprofil — Verlust gegenüber dem eigenen schnellsten Abschnitt',
                   'Sekunden langsamer', [r for r, _ in prof_zeilen], kopf2, 2, 10,
                   beschrifte([rid for _, rid in prof_zeilen]), hoehe=8.6, breite=26)
        c2.x_axis.title = 'Abschnitt'
        gr.add_chart(c2, 'A20')

    gr.page_setup.orientation = 'landscape'
    gr.page_setup.fitToWidth = 1
    gr.page_setup.fitToHeight = 1
    gr.sheet_properties.pageSetUpPr.fitToPage = True
    gr.print_area = 'A1:N38'
    kv.sheet_state = 'hidden'


    ws.freeze_panes = f'{L(C_SEG0)}9'
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = '7:8'
    ws.print_area = f'A1:{L(BREIT)}{letzte}'

    # Rahmen/Faellung der PB-Zeile direkt im XML setzen (siehe
    # markiere_bestzeit()/xlsx_cache.py: openpyxl selbst verwirft das beim
    # Speichern zuverlaessig bei so vielen verschmolzenen Zellen im Blatt).
    # Danach fehlende apply*-Attribute nachtragen (siehe
    # stelle_apply_flags_sicher): ohne applyFill zeigt Excel z.B. die
    # hellgraue/-blaue Schrittzeile teils nicht an, obwohl der Fuellwert
    # korrekt gespeichert ist.
    puffer = io.BytesIO()
    wb.save(puffer)
    fertig = puffer.getvalue()
    if pb_rahmen:
        fertig = setze_rahmen_und_fuellung(fertig, 'Saison', pb_rahmen)
    fertig = stelle_apply_flags_sicher(fertig)
    if hasattr(ziel, 'write'):
        ziel.write(fertig)
    else:
        with open(ziel, 'wb') as f:
            f.write(fertig)
    return ziel, len(lauf), len(vgl)


if __name__ == '__main__':
    athlet = sys.argv[1] if len(sys.argv) > 1 else 'Lars'
    saison = int(sys.argv[2]) if len(sys.argv) > 2 else 2026
    quelle = sys.argv[3] if len(sys.argv) > 3 else 'data/master.csv'
    ziel = f'{athlet}_400mH_{saison}.xlsx'
    p, n, v = baue(load_master(quelle), athlet, saison, ziel)
    print(f'{p}  ({n} Saisonrennen, {v} Vergleichsrennen)')
